# build_oncheon_trees_utf8.py
# 온천공원 XLSX → UTF-8 HTML 재생성기 (안전: 새 폴더에만 출력)

from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re
import html

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
PARK_NAME = "온천공원"

# 입력 XLSX (둘 다 있으면 합쳐서 생성)
XLSX_FILES = [
    ROOT / "parks" / "oncheonpark" / "Oncheon1_trees.xlsx",
    ROOT / "parks" / "oncheonpark" / "Oncheon2_trees.xlsx",
]

OUT_DIR = ROOT / "parks" / "oncheonpark" / "trees_utf8_fixed"

def norm(s: str) -> str:
    return re.sub(r"\s+", "", str(s or "")).strip().lower()

def find_col(headers: List[str], keywords: List[str]) -> Optional[int]:
    nheaders = [norm(h) for h in headers]
    for i, h in enumerate(nheaders):
        for kw in keywords:
            if kw in h:
                return i
    return None

def safe_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def guess_columns(headers):
    # 온천공원 엑셀은 수목코드가 8번 컬럼(0-based index 7)으로 고정
    code_idx = 7  # ← 핵심 수정

    # 수종명
    name_idx = find_col(headers, ["수종명", "수종", "나무", "species", "tree", "name"])

    # 공원명(있으면)
    park_idx = find_col(headers, ["공원", "공원명", "공원코드", "park"])

    return code_idx, name_idx, park_idx

def html_page(code: str, tree_name: str, park_name: str) -> str:
    # DT0001과 비슷한 “기본 상세페이지” 느낌으로 가되, 한글 깨짐 방지(UTF-8 확정)
    title = f"{code} · {tree_name} · {park_name} · THE FORIM".replace("  ", " ").strip()
    title = html.escape(title)

    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title}</title>
  <link rel="stylesheet" href="/assets/style.css" />
</head>
<body>
  <header class="site-header">
    <div class="logo-mark">THE FORIM</div>
    <div class="brand-text">
      <h1>수목 상세 정보</h1>
      <p>조사번호: {html.escape(code)} · {html.escape(park_name)}</p>
    </div>
  </header>

  <main class="container">
    <nav style="margin: 1rem 0 1.5rem; font-size: 0.9rem;">
      <a href="/">메인</a> ·
      <a href="/parks/oncheonpark/">온천공원 수목목록</a>
    </nav>

    <section class="card">
      <h2 style="margin: 0 0 0.75rem;">{html.escape(tree_name)}</h2>

      <div class="kv">
        <div class="kv-row"><div class="k">공원</div><div class="v">{html.escape(park_name)}</div></div>
        <div class="kv-row"><div class="k">수목코드</div><div class="v">{html.escape(code)}</div></div>
      </div>

      <p style="margin-top:1rem; opacity:.8;">
        현장 기록은 아래 버튼으로 입력합니다.
      </p>

      <!-- 기록 버튼(필요하면 나중에 prefill로 교체) -->
      <a class="btn primary" href="/t/{html.escape(code)}">조사 기록</a>
    </section>

    <footer class="site-footer">
      THE FORIM · https://theforim.com
    </footer>
  </main>
</body>
</html>
"""

def build_from_xlsx(xlsx_path: Path) -> List[Tuple[str, str, str]]:
    if not xlsx_path.exists():
        return []

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [safe_text(c).strip() for c in rows[0]]

    def must_idx(colname: str) -> int:
        try:
            return headers.index(colname)
        except ValueError:
            raise RuntimeError(f"Missing column: {colname} | headers={headers}")

    code_idx = must_idx("수목코드")
    name_idx = must_idx("수종명")
    park_idx = must_idx("공원명")

    out: List[Tuple[str, str, str]] = []

    for r in rows[1:]:
        code = safe_text(r[code_idx])
        name = safe_text(r[name_idx])
        park = safe_text(r[park_idx])
        if code and name:
            out.append((code, name, park))

    return out

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    items: List[Tuple[str, str, str]] = []
    for x in XLSX_FILES:
        items.extend(build_from_xlsx(x))

    # 코드 중복 제거(뒤에 온 파일이 우선)
    dedup: Dict[str, Tuple[str, str, str]] = {}
    for code, name, park in items:
        dedup[code] = (code, name, park)

    if not dedup:
        print("No rows found. Check XLSX path or headers.")
        return

    # 페이지 생성
    for code, name, park in dedup.values():
        out_path = OUT_DIR / f"{code}.html"
        out_path.write_text(html_page(code, name, park), encoding="utf-8", newline="\n")

    # index 생성
    links = "\n".join(
        f'      <li><a href="/trees/{html.escape(code)}.html">{html.escape(code)} · {html.escape(name)}</a></li>'
        for code, name, _park in sorted(dedup.values(), key=lambda x: x[0])
    )

    index_html = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{PARK_NAME} · 수목 목록 · THE FORIM</title>
  <link rel="stylesheet" href="/assets/style.css" />
</head>
<body>
  <header class="site-header">
    <div class="logo-mark">THE FORIM</div>
    <div class="brand-text">
      <h1>{PARK_NAME} 수목 목록</h1>
      <p>총 {len(dedup)}건</p>
    </div>
  </header>

  <main class="container">
    <nav style="margin: 1rem 0 1.5rem; font-size: 0.9rem;">
      <a href="/">메인</a> ·
      <a href="/parks/oncheonpark/">온천공원</a>
    </nav>

    <section class="card">
      <ul style="margin:0; padding-left:1.2rem;">
{links}
      </ul>
    </section>

    <footer class="site-footer">
      THE FORIM · https://theforim.com
    </footer>
  </main>
</body>
</html>
"""
    (OUT_DIR / "index.html").write_text(index_html, encoding="utf-8", newline="\n")

    print(f"OK: wrote {len(dedup)} pages to {OUT_DIR}")

if __name__ == "__main__":
    main()